from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import os
import base64
import requests
from PIL import Image
from io import BytesIO
import concurrent.futures
import time
import os
import openpyxl

def scroll_page(driver, scrolls=5, delay=1):
    """滾動頁面指定次數，每次延遲指定秒數"""
    for _ in range(scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(delay)

def verify_image_size(image_data, min_width=150, min_height=150):
    """驗證圖片尺寸是否符合最小要求"""
    try:
        # 使用 PIL 打開圖片
        image = Image.open(BytesIO(image_data))
        width, height = image.size
        return width >= min_width and height >= min_height
    except Exception:
        return False

def search_and_download(food_name, min_width=150, min_height=150):
    print(f"正在搜尋 {food_name} 的圖片...")
    # 設定無頭模式
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--lang=zh-TW")
    
    # 使用 WebDriver Manager 自動配置 ChromeDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    try:
        # 搜尋大尺寸圖片
        search_url = f"https://www.google.com.tw/search?q={food_name}&tbm=isch&tbs=isz:l"
        driver.get(search_url)
        time.sleep(3)  # 等待初始頁面加載

        # 滾動頁面5次，每次間隔1秒
        scroll_page(driver, scrolls=10, delay=2)

        # 解析頁面
        soup = BeautifulSoup(driver.page_source, "lxml")
        results = soup.find_all("img", {"class": "YQ4gaf"})
        image_links = [result.get("src") or result.get("data-src") for result in results if result.get("src") or result.get("data-src")]
        
        print(f"{food_name} 共找到 {len(image_links)} 張圖片")

        # 建立資料夾
        output_dir = os.path.join("images", food_name)
        os.makedirs(output_dir, exist_ok=True)

        downloaded_count = 0
        for index, link in enumerate(image_links):
            try:
                if link.startswith("data:image"):  # Base64 格式圖片
                    header, encoded = link.split(",", 1)
                    image_data = base64.b64decode(encoded)
                else:  # URL 格式圖片
                    response = requests.get(link, timeout=10)
                    image_data = response.content

                # 驗證圖片尺寸
                if verify_image_size(image_data, min_width, min_height):
                    output_path = os.path.join(output_dir, f"{food_name}_{downloaded_count + 1}.jpg")
                    with open(output_path, "wb") as file:
                        file.write(image_data)
                    downloaded_count += 1

            except Exception as e:
                continue

        # 個別食物清冊
        file_name = f"{food_name}_清冊.xlsx"
        sheet_name = "食物清冊"
        first_column = "食物"
        second_column = "檔名"

        file_path = os.path.join(output_dir, file_name)
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            sheet["A1"] = first_column
            sheet["B1"] = second_column
            wb.save(file_path)
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet_values = [sheet.cell(row, 2).value for row in range(2, sheet.max_row+1)]  # 已存在的檔名
        for i in range(1, downloaded_count+1):
            img_name = f"{food_name}_{i}.jpg"
            if img_name in sheet_values: continue
            sheet.append([food_name, img_name])
        wb.save(file_path)

        # 總清冊
        file_name = "總清冊.xlsx"
        sheet_name = "食物清冊"
        first_column = "食物"
        second_column = "檔名"
        file_path = os.path.join("images", file_name)
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            sheet["A1"] = first_column
            sheet["B1"] = second_column
            wb.save(file_path)
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet_values = [sheet.cell(row, 2).value for row in range(2, sheet.max_row+1)]  # 已存在的檔名
        for i in range(1, downloaded_count+1):
            # 判斷是否已經存在
            img_name = f"{food_name}_{i}.jpg"
            if img_name in sheet_values: continue
            sheet.append([food_name, img_name])
        wb.save(file_path)

        print(f"{food_name} 共成功下載 {downloaded_count} 張符合尺寸要求的圖片")

    finally:
        driver.quit()

if __name__ == "__main__":
    file = "key_word_list.txt"
    if not os.path.exists(file):
        with open(file, "w", encoding="utf-8") as f:
            f.write("")
    else:
        # 讀取 key_word_list.txt 並解析
        with open("key_word_list.txt", "r", encoding="utf-8") as file:
            food_list = [food.strip() for food in file.read().split(",")]

        # 設定最小圖片尺寸要求
        MIN_WIDTH = 150
        MIN_HEIGHT = 150

        # 平行執行搜尋與下載
        with concurrent.futures.ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
            futures = [executor.submit(search_and_download, food, MIN_WIDTH, MIN_HEIGHT) 
                    for food in food_list]
            concurrent.futures.wait(futures)